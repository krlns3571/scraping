# coding: utf-8
import datetime
import json
import os
import re
import time

import chromedriver_autoinstaller
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from tqdm import tqdm

chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  # 크롬드라이버 버전 확인
path = chromedriver_autoinstaller.install(True)
file_datetime = datetime.datetime.now().strftime('%y%m%d_%H%M')
os.mkdir(f'./{file_datetime}')
os.mkdir(f'./{file_datetime}/Top 100')


def explicitly_wait(driver, by, name):
    try:
        return WebDriverWait(driver, WAITTIME).until(EC.presence_of_element_located((by, name)))
    except Exception as e:
        raise ValueError(by, name)


def log_flush(driver):
    while driver.get_log('performance'):
        time.sleep(.5)


def print_xlsx(list_, path, header, header_size):
    df = pd.DataFrame(list_, )
    writer = pd.ExcelWriter(f"./{file_datetime}/{path}.xlsx", engine='xlsxwriter', )
    df.to_excel(writer, header=header, index=False)
    for column, column_length in zip(df, header_size):
        col_idx = df.columns.get_loc(column)
        writer.sheets['Sheet1'].set_column(col_idx, col_idx, column_length)
    writer.close()


def log_filter(log_, filter_url):
    return (
        # is an actual response
            log_["method"] == "Network.responseReceived"
            # and json
            and "json" in log_["params"]["response"]["mimeType"]
            # and log_["params"]["response"]["url"].find(filter_url) > 1
            and re.compile(fr'({filter_url})').findall(log_["params"]["response"]["url"])
    )
    # if res:
    #     re.compile(fr'{filter_url}').findall(log_["params"]["response"]["url"])
    #
    # return res


# log_["params"]["response"]["url"].find(filter_url) > 1


def extract_logs(driver, filter_url):
    try:
        cnt = 0
        requests_ids = [[]] * len(filter_url)
        browser_log = []
        results = [[]] * len(filter_url)
        while True:
            [browser_log.append(x) for x in driver.get_log('performance')]
            logs = [json.loads(lr["message"])["message"] for lr in browser_log]

            for idx, url in enumerate(filter_url):
                results[idx] = list(filter(lambda x: log_filter(x, url), logs))

            if sum(x != [] for x in results) == len(filter_url):
                break
            time.sleep(.5)
            cnt += 1
            if cnt > 50:
                # print(browser_log)
                break
                # raise ValueError('no logs')

        for idx, log in enumerate(results):
            # request_id = log["params"]["requestId"]
            try:
                requests_ids[idx] = log[0]["params"]["requestId"]
            except:
                requests_ids[idx] = log
        return requests_ids
    except:
        pass


# pymysql.install_as_MySQLdb()

# warnings.simplefilter("ignore", category=pymysql.Warning)

# Base = declarative_base()
# metadata = Base.metadata

user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36'

WAITTIME = 30
if os.name == 'nt':
    CHROMEDRIVERPATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'chromedriver')
else:
    CHROMEDRIVERPATH = 'chromedriver'

options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--window-size=1920,1080")
options.add_argument("--disable-gpu")
options.add_argument('--incognito')
options.add_argument('--headless')
options.add_argument('--no-sandbox')
options.add_argument('--disable-dev-shm-usage')
options.add_argument("--disable-setuid-sandbox")
options.add_argument("--lang=en-US")

options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_experimental_option('useAutomationExtension', False)
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument('user-agent={0}'.format(user_agent))

prefs = {'disk-cache-size': 40960}
options.add_experimental_option("prefs", prefs)
caps = DesiredCapabilities.CHROME
caps['goog:loggingPrefs'] = {'performance': 'ALL'}
s = Service(path)

infos = []
no_data = []
dates = ['24h', '7d', '30d', '3M', '1y']


def driver_setting():
    driver = webdriver.Chrome(service=s, options=options,
                              desired_capabilities=caps)
    driver.execute_cdp_cmd('Network.setUserAgentOverride', {
        "userAgent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.53 Safari/537.36'})
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver


if __name__ == '__main__':
    driver = driver_setting()
    driver.get('https://nftgo.io/analytics/market-overview')
    print('https://nftgo.io/analytics/market-overview')
    while True:
        try:
            if driver.find_element(By.XPATH, "//span[contains(@class,'overview_emp__ImpDY')]").text == '0':
                time.sleep(1)
                continue
            else:
                break
        except:
            pass

    all_buttons = driver.find_elements(By.XPATH, "//div[contains(text(),'All')][contains(@class,'time-span')]")
    all_buttons[0].click()
    time.sleep(.5)
    cnt = 3
    while cnt:
        try:
            time.sleep(.5)
            all_buttons[1].click()
            driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
            time.sleep(.5)
            all_buttons[2].click()
            break
        except:
            cnt -=1
            pass
    time.sleep(2)
    marketcap, metrics, collections, holder_trader = extract_logs(driver,
                                                                  ['marketcap-volume\?range=all', 'category\/metrics',
                                                                   "pie\/collections", "holder-trader\?range=all"])
    marketcap_data = json.loads(driver.execute_cdp_cmd(
        "Network.getResponseBody", {"requestId": marketcap})['body'])['data']
    metrics_data = json.loads(driver.execute_cdp_cmd(
        "Network.getResponseBody", {"requestId": metrics})['body'])['data']
    collections_data = json.loads(driver.execute_cdp_cmd(
        "Network.getResponseBody", {"requestId": collections})['body'])['data']
    holder_trader_data = json.loads(driver.execute_cdp_cmd(
        "Network.getResponseBody", {"requestId": holder_trader})['body'])['data']
    time.sleep(2)

    marketcap_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                           marketcap_data['marketCap']['values']['x']]
    marketcap_values = [round(x, 2) for x in marketcap_data['marketCap']['values']['y']]
    volume_values = [round(x, 2) for x in marketcap_data['volume']['values']['y']]
    holders_values = [round(x, 2) for x in holder_trader_data['holders']['values']['y']]
    traders_values = [round(x, 2) for x in holder_trader_data['traders']['values']['y']]
    # pd.DataFrame(list(zip(marketcap_timestamp, marketcap_values, volume_values, holders_values, traders_values)))
    print_xlsx(
        pd.DataFrame(list(zip(marketcap_timestamp, marketcap_values, volume_values, holders_values, traders_values))),
        "Market Cap Volume_Holders Traders", ['Date', 'Market Cap', 'Volume', 'Holder', 'Traders'],
        [20, 20, 20, 20, 20]
    )
    print('Market Cap Volume_Holders Traders 저장완료')
    print_xlsx(
        pd.DataFrame([round(x, 2) for x in metrics_data['marketCap']['values']['y']]).T,
        "Category Market Cap_Market Cap", metrics_data['marketCap']['values']['x'],
        [15 for _ in metrics_data['marketCap']['values']['x']]
    )
    print('Category Market Cap_Market Cap 저장완료')
    # market_cap_handles = driver.find_elements(By.XPATH, "//div[contains(@class,'market-cap_handle')]//span")
    #
    # market_cap_handles[2].click()
    # log_flush(driver)
    # driver.find_element(By.XPATH, "//div[contains(@class,'category-market-cap_timeSelect')]").click()
    #
    # driver.find_element(By.XPATH, "//div[contains(@class,'time-span')]/span[contains(text(),'24H')]").click()
    cmc_volume = []
    for date in dates:
        cmc_volume.append([date] + [round(x, 2) for x in metrics_data[f'volume{date}']['values']['y']])
    cmc_volume.append(["ALL"] + [round(x, 2) for x in metrics_data[f'volume']['values']['y']])

    print_xlsx(
        pd.DataFrame(cmc_volume),
        "Category Market Cap_Volume", [x for x in ["Duration"] + metrics_data[f'volume']['values']['x']],
        [15 for _ in [x for x in ["Duration"] + metrics_data[f'volume']['values']['x']]]
    )
    print('Category Market Cap_Volume 저장완료')
    cmc_liquidity = []
    for date in dates:
        cmc_liquidity.append(
            [date] + [f"{round(x * 100, 2)}%" for x in metrics_data[f'liquidity{date}']['values']['y']])

    cmc_liquidity.append(["ALL"] + [f"{round(x * 100, 2)}%" for x in metrics_data[f'liquidity']['values']['y']])

    print_xlsx(
        pd.DataFrame(cmc_liquidity),
        "Category Market Cap_Liquidity", [x for x in ['Duration'] + metrics_data[f'volume']['values']['x']],
        [15 for _ in [x for x in ['Duration'] + metrics_data[f'volume']['values']['x']]]
    )
    print('Category Market Cap_Liquidity 저장완료')
    print_xlsx(
        pd.DataFrame(
            [f'{round(x["ratio"] * 100, 2)}%' for x in collections_data['marketCap']]).T,
        "Collection Distribution_Market Cap",
        [x['collection']['name'] for x in collections_data['marketCap'] if x['collection'] is not None] + ['Others'],
        [15 for _ in [x for x in collections_data['marketCap'] if x['collection'] is not None] + ['Others']]
    )
    print('Collection Distribution_Market Cap 저장완료')
    print_xlsx(
        pd.DataFrame(
            [f'{round(x["value"], 2)}' for x in collections_data['marketCap']]).T,
        "Top Collections_Market Cap",
        [x['collection']['name'] for x in collections_data['marketCap'] if x['collection'] is not None] + ['Others'],
        [15 for _ in [x for x in collections_data['marketCap'] if x['collection'] is not None] + ['Others']]
    )
    print('Top Collections_Market Cap 저장완료')
    cd_volume = []
    cd_volume2 = []
    for date in dates:
        cd_volume_names = [x['collection']['name'] for x in collections_data[f'volume{date}'] if
                           x['collection'] is not None] + ['Others']
        cd_volume.append(
            [date] + cd_volume_names)
        cd_volume2.append(
            [date] + cd_volume_names)
        cd_volume_ratio = [
            # str(round(x['value'], 2)) +
            f"{round(x['ratio'] * 100, 2)}%" for x in
            collections_data[f'volume{date}']]
        cd_volume_value = [str(round(x['value'], 2)) for x in collections_data[f'volume{date}']]
        cd_volume.append(
            [date] + cd_volume_ratio)
        cd_volume2.append(
            [date] + cd_volume_value)

    cd_volume_names = [x['collection']['name'] for x in collections_data[f'volume'] if
                       x['collection'] is not None] + ['Others']
    cd_volume.append(["ALL"] + cd_volume_names)
    cd_volume2.append(["ALL"] + cd_volume_names)
    cd_volume_ratio = [
        # str(round(x['value'], 2)) +
        f"{round(x['ratio'] * 100, 2)}%" for x in
        collections_data[f'volume']]
    cd_volume_value = [str(round(x['value'], 2)) for x in collections_data[f'volume']]
    cd_volume.append(["ALL"] + cd_volume_ratio)
    cd_volume2.append(["ALL"] + cd_volume_value)

    print_xlsx(
        pd.DataFrame(
            cd_volume),
        "Collection Distribution_Volume",
        ['Duration'] + [f'Project {x + 1}' for x, _ in enumerate(cd_volume_names)] + ['Others'],
        [9, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
    )
    print('Collection Distribution_Volume 저장완료')
    print_xlsx(
        pd.DataFrame(
            cd_volume2),
        "Top Collections_Volume",
        ['Duration'] + [f'Project {x + 1}' for x, _ in enumerate(cd_volume_names)] + ['Others'],
        [9, 25, 25, 25, 25, 25, 25, 25, 25, 25, 25]
    )
    print('Top Collections_Volume 저장완료')
    for filename in ["Collection Distribution_Volume", "Top Collections_Volume"]:
        wb = openpyxl.load_workbook(f"{os.path.abspath(f'./{file_datetime}/{filename}.xlsx')}")
        ws = wb.active

        ws.merge_cells('A2:A3')
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A4:A5')
        ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A6:A7')
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A8:A9')
        ws['A8'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A10:A11')
        ws['A10'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A12:A13')
        ws['A12'].alignment = Alignment(horizontal='center', vertical='center')
        wb.save(f"{os.path.abspath(f'./{file_datetime}/{filename}.xlsx')}")
    # cd_volume.append(["ALL"] + [f"{round(x * 100, 3)}%" for x in metrics_data[f'liquidity']['values']['y']])

    driver.close()

    cd_volume = None
    cd_volume2 = None
    cd_volume_names = None
    cd_volume_ratio = None
    cd_volume_value = None
    cmc_liquidity = None
    cmc_volume = None
    collections_data = None
    cmc_liquidity = None
    cmc_volume = None
    holder_trader_data = None
    holders_values = None
    marketcap_data = None
    marketcap_timestamp = None
    marketcap_values = None
    metrics_data = None
    traders_values = None
    volume_values = None

    driver = driver_setting()

    driver.get("https://nftgo.io/analytics/top-collections?tag=all&timeSpan=30d&sort=marketCap&rarity=false")
    print('https://nftgo.io/analytics/top-collections 페이지 접근')
    collections = []
    market_caps = []
    volumes = []
    floor_price = []
    whales = []

    flag = True
    while flag:
        while True:
            if len(driver.find_elements(By.XPATH,
                                        "//div[contains(@class,'rank-list-item_rankItem')][contains(@id,'fixed')]")) == 50:
                break
            else:
                time.sleep(1)
                continue

        [collections.append(x.text) for x in
         driver.find_elements(By.XPATH, '//div[contains(@id,"fixed-")]/div[2]/div/div[4]')]
        [market_caps.append(x.text.replace('$', '').replace(',', '')) for x in
         driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[1]/div')]
        [volumes.append(x.text.replace('$', '').replace(',', '')) for x in
         driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[3]/div')]
        [floor_price.append(x.text) for x in driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[5]')]
        [whales.append(x.text) for x in driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[7]/div/div')]

        if flag:
            driver.find_element(By.XPATH, '//body').send_keys(Keys.END)
            time.sleep(1)
            driver.find_element(By.XPATH, "//div[text()=2][contains(@class,'pagination_rank')]").click()
            flag = False
        else:
            break
    [collections.append(x.text) for x in
     driver.find_elements(By.XPATH, '//div[contains(@id,"fixed-")]/div[2]/div/div[4]')]
    [market_caps.append(x.text.replace('$', '').replace(',', '')) for x in
     driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[1]/div')]
    [volumes.append(x.text.replace('$', '').replace(',', '')) for x in
     driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[3]/div')]
    [floor_price.append(x.text) for x in driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[5]')]
    [whales.append(x.text) for x in driver.find_elements(By.XPATH, '//div[contains(@id,"item-")]/div[7]/div/div')]

    print_xlsx(
        pd.DataFrame(
            [list(range(1, 101)), collections, market_caps, volumes, floor_price, whales]).T,
        "Top 100/Top100_NFT Collections",
        ['rank', 'Project Title', 'Market Cap', 'Volume', 'Floor Price', 'Whale'],
        [9, 20, 20, 20, 20, 20, 20, ]
    )
    print('Top 100/Top100_NFT Collections 저장완료')
    driver.execute_script(
        'window.open("https://api.nftgo.io/api/v1/ranking/collections?offset=0&limit=100&by=marketCap&interval=30d&asc=-1&rarity=-1&fields=marketCap,marketCapChange30d,relMarketCap,buyerNum30d,buyerNum30dChange30d,sellerNum30d,sellerNum30dChange30d,liquidity30d,liquidity30dChange30d,saleNum30d,saleNum30dChange30d,volume30d,volume30dChange30d,relVolume30d,traderNum30d,traderNum30dChange30d,holderNum,holderNumChange30d,whaleNum,whaleNumChange30d,orderAvgPriceETH30d,orderAvgPriceETH30dChange30d,orderAvgPrice30d,orderAvgPrice30dChange30d,floorPrice,floorPriceChange30d,relMaxFloorPrice,relMaxOrderAvgPrice30d,relMaxLiquidity30d,relMaxSaleNum,relMaxSaleNum30d,relMaxBuyerNum,relMaxSellerNum,relMaxWhaleNum,relMaxTraderNum,relMaxHolderNum,relMaxTraderNum30d,relMarketCap,relMaxBuyerNum30d,relMaxSellerNum30d");')
    driver.switch_to.window(driver.window_handles[1])
    slugs = [x['slug'] for x in json.loads(driver.find_element(By.XPATH, '//body').text)['data']['list']]
    print('Top 100 list 추출 완료 // Top100 수집시작')
    driver.quit()
    rank = 0
    for slug in tqdm(slugs, unit='collection'):
        cnt = 3
        while cnt:
            try:
                driver = driver_setting()
                driver.get(f"https://nftgo.io/collection/{slug}/overview")

                rank += 1

                explicitly_wait(driver, By.XPATH, "//div[contains(@class,'volume-graph_legendGroup')]")
                time.sleep(2)
                # print(driver.current_url)

                for x in driver.find_elements(By.XPATH, "//div[contains(text(),'All')][contains(@class,'time-span')]")[
                         :2]:
                    driver.execute_script("arguments[0].scrollIntoView();", x)

                    driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_UP)
                    time.sleep(.5)
                    x.click()
                while True:
                    try:
                        explicitly_wait(driver, By.XPATH,
                                        "//div[contains(@class,'price-graph')]//div[contains(@class,'legends_legend')]")
                        time.sleep(3)
                        break
                    except:
                        pass
                driver.find_element(By.XPATH, "//div[contains(@class,'timeSpanSelect')]").click()
                driver.find_element(By.XPATH,
                                    "//div[contains(@class,'select')]//span[contains(text(),'All')]/..").click()

                driver.find_element(By.XPATH, "//body").send_keys(Keys.HOME)
                time.sleep(.3)
                driver.find_element(By.XPATH, "//span[contains(text(),'Holder')]/..").click()
                time.sleep(.3)

                for x in driver.find_elements(By.XPATH, "//div[contains(text(),'All')][contains(@class,'time')]"):
                    x.click()
                    driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_DOWN)
                    time.sleep(.5)
                driver.find_element(By.XPATH, "//body").send_keys(Keys.PAGE_UP)
                time.sleep(1)
                driver.find_element(By.XPATH, "//div[contains(text(),'Amount')]").click()

                time.sleep(2)
                a, b, c, d, e, f, g, h = extract_logs(driver, ["marketcap-volume.*&range=all", "price.*&range=all",
                                                            "holders-chart.*&range=all",
                                                            "coll-nft-held-time.*&range=all", "holdingTimeDis",
                                                            "holdingAmountDis", "coll-holder-nft-num.*&range=all", "transfer.*&range=all"])
                marketcap_volume_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": a})['body'])['data']
                price_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": b})['body'])['data']
                holders_chart_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": c})['body'])['data']
                concentration_value_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": d})['body'])['data']
                concentration_nft_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": g})['body'])['data']
                holdingtimedis_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": e})['body'])['data']
                holdingamountdis_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": f})['body'])['data']
                trans_liquid_data = json.loads(driver.execute_cdp_cmd(
                    "Network.getResponseBody", {"requestId": h})['body'])['data']
                break
            except Exception as e:
                # driver.quit()
                if cnt == 1:
                    break
                else:
                    cnt -= 1
                    driver.close()
                    pass
        try:
            marketcap_volume_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": a})['body'])['data']
        except:
            pass
        try:
            price_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": b})['body'])['data']
        except:
            pass
        try:
            holders_chart_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": c})['body'])['data']
        except:
            pass
        try:
            concentration_value_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": d})['body'])['data']
        except:
            pass
        try:
            concentration_nft_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": g})['body'])['data']
        except:
            pass
        try:
            holdingtimedis_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": e})['body'])['data']
        except:
            pass
        try:
            holdingamountdis_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": f})['body'])['data']
        except:
            pass
        try:
            trans_liquid_data = json.loads(driver.execute_cdp_cmd(
                "Network.getResponseBody", {"requestId": h})['body'])['data']
        except:
            pass
        collection_name = driver.find_element(By.XPATH, "//span[contains(@class,'collectionName')]").text.replace(":",
                                                                                                                  "")
        total_supply = driver.find_element(By.XPATH, "//span[contains(@class,'nftNum')]").text.replace(',', '')
        os.mkdir(f'./{file_datetime}/Top 100/{rank}_{collection_name}')

        try:
            overview_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                                  marketcap_volume_data['marketCap']['values']['x']]
            print_xlsx(
                pd.DataFrame(
                    [overview_timestamp, [collection_name for _ in range(0, len(overview_timestamp))],
                     [total_supply for _ in range(0, len(overview_timestamp))],
                     marketcap_volume_data['marketCap']['values']['y'],
                     marketcap_volume_data['volume']['values']['y'], ]
                ).T,
                f"Top 100/{rank}_{collection_name}/OverView_Marketcap_Volume",
                ['Date', 'Project Title', 'Total Supply', 'Market Cap', 'Volume', ],
                [12, 20, 20, 20, 20, 20, ]
            )
        except:
            print(f"{rank}_{collection_name}/OverView_Marketcap_Volume")
            # overview_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
            #                       marketcap_volume_data['marketCap']['values']['x']]
            # print_xlsx(
            #     pd.DataFrame(
            #         [overview_timestamp, [collection_name for _ in range(0, len(overview_timestamp))],
            #          [total_supply for _ in range(0, len(overview_timestamp))],
            #          marketcap_volume_data['marketCap']['values']['y'],
            #          marketcap_volume_data['volume']['values']['y'], ]
            #     ).T,
            #     f"Top 100/{rank}_{collection_name}/OverView",
            #     ['Date', 'Project Title', 'Total Supply', 'Market Cap', 'Volume', ],
            #     [12, 20, 20, 20, 20, 20, ]
            # )
        try:
            price_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                                  price_data['avgPrice']['values']['x']]
            print_xlsx(
                pd.DataFrame(
                    [price_timestamp,
                     price_data['floorPrice']['values']['y'], price_data['avgPrice']['values']['y'] ]
                ).T,
                f"Top 100/{rank}_{collection_name}/OverView_Price",
                ['Date','Floor Price', 'Average Price' ],
                [12, 20, 20, ]
            )
            #OverView_price
            # price_data['floorPrice']['values']['y'], price_data['avgPrice']['values']['y']
        except:
            print(f"{rank}_{collection_name}/OverView_Price")
        try:
            trans_liquid_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for
                                       x in trans_liquid_data['sales']['values']['x']]

            print_xlsx(
                pd.DataFrame([trans_liquid_timestamp, trans_liquid_data['sales']['values']['y']]).T,
                f"Top 100/{rank}_{collection_name}/Sales",
                ['Date', 'Sales'],
                [20, 20]
            )
        except:
            print(f"{rank}_{collection_name}/Sales")
            pass
        try:
            holder_chart_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                                      holders_chart_data['holderChart']['values']['x']]

            print_xlsx(
                pd.DataFrame([holder_chart_timestamp, holders_chart_data['holderChart']['values']['y'],
                              [str(x) if x is not None else '0' for x in
                               holders_chart_data['whaleChart']['values']['y']],
                              [str(x) if x is not None else '0' for x in
                               holders_chart_data['blueChipHolderChart']['values']['y']]]).T,
                f"Top 100/{rank}_{collection_name}/Holder Trends",
                ['Date', 'Holders', 'Whales','blueChipHolder'],
                [12, 20, 20,20]
            )
        except:
            print(f"{rank}_{collection_name}/Holder Trends")
            pass

        try:
            concentration_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                                       concentration_value_data['chart1']['values']['x']]

            print_xlsx(
                pd.DataFrame([concentration_timestamp,
                              concentration_value_data['chart1']['values']['y'],
                              concentration_value_data['chart1to7']['values']['y'],
                              concentration_value_data['chart7to30']['values']['y'],
                              concentration_value_data['chart30to90']['values']['y'],
                              concentration_value_data['chart90to365']['values']['y'],
                              concentration_value_data['chart365']['values']['y']]).T,
                f"Top 100/{rank}_{collection_name}/Holding Period Trends",
                ['Date', '<24H', '1-7D', '7-30D', '30D-3M', '3M-1Y', '<1Y'],
                [12, 20, 20, 20, 20, 20, 20]
            )
        except:
            print(f"{rank}_{collection_name}/Holding Period Trends")
            pass

        try:
            concentration_timestamp = [str(datetime.datetime.fromtimestamp(x / 1000).date()) for x in
                                       concentration_nft_data['chart1']['values']['x']]

            print_xlsx(
                pd.DataFrame([concentration_timestamp,
                              concentration_nft_data['chart1']['values']['y'],
                              concentration_nft_data['chart1to3']['values']['y'],
                              concentration_nft_data['chart3to10']['values']['y'],
                              concentration_nft_data['chart10to50']['values']['y'],
                              concentration_nft_data['chart50to100']['values']['y'],
                              concentration_nft_data['chart100']['values']['y']]).T,
                f"Top 100/{rank}_{collection_name}/Holding Amount Trends",
                ['Date', '1', '2-3', '4-10', '11-50', '51-100', '>100'],
                [12, 20, 20, 20, 20, 20, 20]
            )
        except:
            print(f"{rank}_{collection_name}/Holding Amount Trends")
            pass
        try:
            print_xlsx(
                pd.DataFrame([f"{round(x * 100, 2)}%" for x in holdingamountdis_data['disPercent']]).T,
                f"Top 100/{rank}_{collection_name}/Holding Amount Distribution",
                ["1", "2-3", "4-10", "11-50", "51-100", ">100"],
                [10, 10, 10, 10, 10, 10]
            )
        except:
            print(f"{rank}_{collection_name}/Holding Amount Distribution")
            pass
        driver.quit()

    # try:
    #     driver.quit()
    #     driver.close()
    # except:
    #     pass
    # driver.close()
    # print(1)
    print(f"{os.path.abspath(f'./{file_datetime}')} 에 모든 결과가 저장되었습니다.\n해당 창은 꺼주셔도 좋습니다.")
    exit(1)
