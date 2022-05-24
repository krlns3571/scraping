#!/usr/bin/env python

import tkinter as tk
import tkinter.font
from tkinter import ttk
from tkinter import filedialog as fd
import tkinter.messagebox as mbox

import datetime, os

import openpyxl

def get_resource_path(filename):
    return os.path.join(os.path.dirname(__file__), filename)

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string        

# MainFrame
class frmMain(tk.Frame):
    def __init__(self, parent):
        tk.Frame.__init__(self, parent)
        self.pack(padx=10, pady=2)
        
        self.defaultFont = tkinter.font.nametofont("TkDefaultFont")
        self.defaultFont.configure(family = "Gulim", size = 10)
        
        parent.geometry("350x350")
        icopath = get_resource_path('favicon.ico')
        if os.path.isfile(icopath):
            parent.iconbitmap(icopath)

        self.frmT1=tk.Frame(parent)
        self.frmT1.pack(side=tk.TOP)
        
        label = tk.Label(self.frmT1, text="대조파일선택", width=11, height=1, pady=3)
        label.pack(side=tk.LEFT)
        self.btnSelectComp = tk.Button(self.frmT1, width=10, text="찾아보기...", command=self.SelectComp)
        self.btnSelectComp.pack(side=tk.LEFT)
        self.entCompFile=tk.Entry(self.frmT1)
        self.entCompFile.pack(side=tk.LEFT, padx=5, pady=3, fill=tk.X)

        self.frmT2=tk.Frame(parent)
        self.frmT2.pack(side=tk.TOP)
        label = tk.Label(self.frmT2, text="마켓파일선택", width=11, height=1, pady=3)
        label.pack(side=tk.LEFT)
        self.btnSelectMarket = tk.Button(self.frmT2, width=10, text="찾아보기...", command=self.SelectMarketX)
        self.btnSelectMarket.pack(side=tk.LEFT)
        self.entMarketFile=tk.Entry(self.frmT2)
        self.entMarketFile.pack(side=tk.LEFT, padx=5, pady=3, fill=tk.X)
        
        self.frmT3=tk.Frame(parent)
        self.frmT3.pack(side=tk.TOP)
        self.btnBack1st = tk.Button(self.frmT3, width=10, text="↑이전단계", state=tk.DISABLED, command=self.Back1st)
        self.btnBack1st.pack(side=tk.LEFT, padx=5, pady=3)
        self.btnGoto2nd = tk.Button(self.frmT3, width=10, text="다음단계↓", command=self.Goto2nd)
        self.btnGoto2nd.pack(side=tk.LEFT, padx=5, pady=3)

        self.frmT4=tk.Frame(parent)
        self.frmT4.pack(side=tk.TOP)
        
        label = tk.Label(self.frmT4, text="마켓 선택: ", height=1, pady=3)
        label.pack(side=tk.LEFT)
        self.cboMarket=tkinter.ttk.Combobox(self.frmT4, height=15, values=[])
        self.cboMarket.pack(side=tk.LEFT)
        
        self.frmT5=tk.Frame(parent)
        self.frmT5.pack(side=tk.TOP)
        label = tk.Label(self.frmT5, text="무게값을 반영할 ID 선택", height=1, pady=3)
        label.pack(side=tk.TOP)
        
        scrollbar=tkinter.Scrollbar(self.frmT5)
        scrollbar.pack(side="right", fill="y")
        
        self.lstIds = tk.Listbox(self.frmT5, selectmode='multiple', height=10, width=40, activestyle=tk.DOTBOX)
        scrollbar["command"]=self.lstIds.yview
        self.lstIds.pack(side=tk.TOP)
        
        self.frmT6=tk.Frame(parent)
        self.frmT6.pack(side=tk.TOP)
        btnSelectAll = tk.Button(self.frmT6, width=9, text="모두선택", command=self.SelectAll)
        btnSelectAll.pack(side=tk.LEFT, padx=5, pady=3)
        btnSelectNone = tk.Button(self.frmT6, width=9, text="모두해제", command=self.SelectNon)
        btnSelectNone.pack(side=tk.LEFT, padx=5, pady=3)
        btnSelectRev = tk.Button(self.frmT6, width=9, text="선택반전", command=self.SelectRev)
        btnSelectRev.pack(side=tk.LEFT, padx=5, pady=3)
        
        self.frmT7=tk.Frame(parent)
        self.frmT7.pack(side=tk.TOP)
        self.btnExec = tk.Button(self.frmT7, width=10, text="▶ 실행", pady=4, state=tk.DISABLED, command=self.RunWeight)
        self.btnExec.pack(side=tk.LEFT, padx=5, pady=5)
        
        
    def SelectAll(self):
        self.lstIds.selection_set(0, tk.END)
        
    def SelectNon(self):
        self.lstIds.selection_clear(0, tk.END)
        
    def SelectRev(self):
        cnt = self.lstIds.size()
        sel_list = list( self.lstIds.curselection() )
        for i in range(cnt):
            if i in sel_list:
                self.lstIds.selection_clear(i)
            else:
                self.lstIds.selection_set(i)
                
    def SelectComp(self):
        filename = fd.askopenfilename(filetypes=[("엑셀 XLSX 파일", ".xlsx"), ("모든 파일", ".*")])
        if filename:
            self.entCompFile.delete(0, tk.END)
            self.entCompFile.insert(0, filename)
            
    def SelectMarketX(self):
        filename = fd.askopenfilename(filetypes=[("엑셀 XLSX 파일", ".xlsx"), ("모든 파일", ".*")])
        if filename:
            self.entMarketFile.delete(0, tk.END)
            self.entMarketFile.insert(0, filename)
                
    def Goto2nd(self):
        if self.entCompFile.get().strip() == '' or self.entMarketFile.get().strip() == '':
            mbox.showwarning("오류", "파일을 선택해 주세요.")
            return
            
        try:
            self.wbComp = openpyxl.load_workbook( self.entCompFile.get() )
            self.wbMarket = openpyxl.load_workbook( self.entMarketFile.get(), data_only=True )
            
            self.cboMarket['values'] = self.wbComp.sheetnames
            id_list = self.wbMarket.sheetnames
            self.lstIds.delete(0, tk.END)
            for ids in id_list:
                # 숨김시트 빼기
                if self.wbMarket[ids].sheet_state == 'hidden': continue
                self.lstIds.insert(tk.END, ids)
            #self.SelectAll()
        except:
            mbox.showwarning("오류", "파일을 처리할 수 없습니다.")
            return
            
        self.btnSelectComp.configure(state=tk.DISABLED)
        self.btnSelectMarket.configure(state=tk.DISABLED)
        self.entCompFile.configure(state=tk.DISABLED)
        self.entMarketFile.configure(state=tk.DISABLED)
        self.btnBack1st.configure(state=tk.ACTIVE)
        self.btnGoto2nd.configure(state=tk.DISABLED)
        self.btnExec.configure(state=tk.ACTIVE)

    def Back1st(self):
        self.btnSelectComp.configure(state=tk.ACTIVE)
        self.btnSelectMarket.configure(state=tk.ACTIVE)
        self.entCompFile.configure(state=tk.NORMAL)
        self.entMarketFile.configure(state=tk.NORMAL)
        self.btnBack1st.configure(state=tk.DISABLED)
        self.btnGoto2nd.configure(state=tk.ACTIVE)
        self.btnExec.configure(state=tk.DISABLED)
        
    def RunWeight(self):
        if self.cboMarket.get().strip == '':
            mbox.showwarning("오류", "마켓을 선택해 주세요.")
            return
        sel_list = list( self.lstIds.curselection() )
        if sel_list == []:
            mbox.showwarning("오류", "적용할 ID를 선택헤 주세요.")
            return
            
        try:
        #if 1:
            market_name = self.cboMarket.get()
            colComp, colId, colSet = 0, 0, 0
            
            # 열찾기
            for i in range(1, 16385):
                if self.wbComp[market_name]["%s2" % colnum_string(i) ].value == '회사':
                    colComp = i
                elif self.wbComp[market_name]["%s2" % colnum_string(i) ].value == '아이디':
                    colId = i
                elif self.wbComp[market_name]["%s2" % colnum_string(i) ].value == '전송세트(템플릿)명':
                    colSet = i
                if colComp and colId and colSet:
                    break

            if colComp == 0 or colId == 0 or colSet == 0:
                mbox.showwarning("오류", "전송세트 엑셀의 해당 마켓 시트의 2번째 행에서 지정된 열 이름이 발견되지 않습니다.")
                return
                    
            row_cnt = self.wbComp[market_name].max_row

            for sid in sel_list:
                shop_id = self.lstIds.get(sid)
                colTmplNm, colWeight = 0, 0
            
                for i in range(1, 16385):
                    if self.wbMarket[shop_id]["%s1" % colnum_string(i) ].value == '템플릿 이름':
                        colTmplNm = i
                    elif self.wbMarket[shop_id]["%s1" % colnum_string(i) ].value == '무게':
                        colWeight = i
                    if colTmplNm and colWeight:
                        break
                if colTmplNm == 0 or colWeight == 0:
                    break
                
                m_row_cnt = self.wbMarket[shop_id].max_row
                weight_list = []
                for i in range(2, m_row_cnt+1):
                    try:
                        if self.wbMarket[shop_id]["%s%d" % (colnum_string(colTmplNm), i)].value.strip() == '': continue
                    except:
                        continue
                    weight_list.append({
                        'weight': self.wbMarket[shop_id]["%s%d" % (colnum_string(colWeight), i)].value,
                        'name':   self.wbMarket[shop_id]["%s%d" % (colnum_string(colTmplNm), i)].value,
                    })
                #print(weight_list)
                
                # 훑기
                for i in range(2, row_cnt+1):
                    try:
                        # 아이디가 동일하지 않거나 공란이면 패스
                        if self.wbComp[market_name]["%s%d" % (colnum_string(colId), i)].value.strip() != shop_id: continue
                        elif self.wbComp[market_name]["%s%d" % (colnum_string(colComp), i)].value.strip() == '': continue
                    except:
                        continue
                    
                    # 텍스트 분할 약부호-ID-무게
                    weight_arr = self.wbComp[market_name]["%s%d" % (colnum_string(colComp), i)].value.strip().split('-')
                    for wl in weight_list:
                        if wl['weight'] == weight_arr[2]:
                            self.wbComp[market_name]["%s%d" % (colnum_string(colSet), i)].value = wl['name']
                            break
            
            # 저장
            now = datetime.datetime.now()
            self.wbComp.save('./out/전송시트_템플릿_출력결과_%s.xlsx' % now.strftime("%Y%m%d%H%M%S") )
            mbox.showinfo("결과", "출력이 완료되었습니다.\n경로: ./out/전송시트_템플릿_출력결과_%s.xlsx" % now.strftime("%Y%m%d%H%M%S") )
            
        #if 0:
        except:
            mbox.showwarning("오류", "작업을 처리할 수 없습니다.")
            pass
        
if __name__ == "__main__":
    root = tk.Tk()
    myapp=frmMain(root)
    root.resizable(False, False)
    root.title('Shop 엑셀파일 자동화 B Prototype')
    #root.eval('tk::PlaceWindow . center')
    root.mainloop()
